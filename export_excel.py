from __future__ import annotations

from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Alignment, Border, Side


def _autosize(ws, df: pd.DataFrame, max_width: int = 60) -> None:
    """
    Bezpieczne autosize: wartości mogą być float/NaN/datetime itd.
    Wszystko liczymy po str().
    """
    for i, col in enumerate(df.columns, start=1):
        s = df[col]

        # bierzemy próbkę, zamieniamy na stringi (NaN -> "")
        sample = s.head(200).astype("string").fillna("")
        lens = [len(str(col))] + [len(v) for v in sample.tolist()]

        w = max(lens) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(w, max_width)


def _format_numbers(ws, df: pd.DataFrame) -> None:
    int_fmt = "0"
    float_fmt = "0.00"

    for col_idx, col_name in enumerate(df.columns, start=1):
        s = df[col_name]
        if pd.api.types.is_integer_dtype(s):
            fmt = int_fmt
        elif pd.api.types.is_float_dtype(s):
            fmt = float_fmt
        else:
            continue

        for row_idx in range(2, len(df) + 2):
            ws.cell(row=row_idx, column=col_idx).number_format = fmt


def _force_comma_text_for_columns(ws, col_names: list[str], decimals: int = 2) -> None:
    """
    Zamienia wartości liczbowe w podanych kolumnach na TEKST z przecinkiem
    i ustawia number_format="@" żeby Excel nie przerabiał z powrotem.
    """
    headers = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    for name in col_names:
        col_idx = headers.get(name)
        if not col_idx:
            continue

        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=col_idx)
            v = cell.value

            # zostaw teksty typu "brak pomiaru"
            if v is None or isinstance(v, str):
                continue

            try:
                s = f"{float(v):.{decimals}f}".replace(".", ",")
                cell.value = s
                cell.number_format = "@"  # tekst
            except Exception:
                pass


# --- opis (żółte tło) ---
_YELLOW = PatternFill("solid", fgColor="FFF200")
_ALIGN = Alignment(vertical="top", horizontal="center", wrap_text=True)
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _add_description_block(
    ws,
    text: str,
    start_cell: str,
    end_cell: str,
    col_width: float = 22.0,
) -> None:
    ws.merge_cells(f"{start_cell}:{end_cell}")
    cell = ws[start_cell]
    cell.value = text
    cell.fill = _YELLOW
    cell.alignment = _ALIGN
    cell.border = _BORDER

    # ustaw szerokość kolumn w zakresie opisu (np. H..N)
    start_col = ord(start_cell[0].upper())
    end_col = ord(end_cell[0].upper())
    for c in range(start_col, end_col + 1):
        ws.column_dimensions[chr(c)].width = col_width


def _write_package_type_share_summary(ws, weighted_avg_len: float, predicted_efficiency: int) -> None:
    ws["I1"].value = f"avg_len: {weighted_avg_len:.2f}".replace(".", ",")
    ws["I2"].value = f"predicted_eff: {predicted_efficiency}"



def write_report_xlsx(
    path: Path,
    sheets: Dict[str, pd.DataFrame],
    sheet_order: Optional[Iterable[str]] = None,
    descriptions: Optional[Dict[str, Tuple[str, str, str]]] = None,
    package_type_share_summary: Optional[Tuple[float, int]] = None,
) -> None:
    """
    descriptions[sheet_name] = (text, start_cell, end_cell)
    sheet_order: wymusza kolejność arkuszy (reszta dopisana na końcu)

    package_type_share_summary = (weighted_avg_len, predicted_efficiency)
    """
    path.parent.mkdir(parents=True, exist_ok=True)

    # kolejność arkuszy
    if sheet_order is None:
        order = list(sheets.keys())
    else:
        order = []
        seen = set()
        for name in sheet_order:
            if name in sheets and name not in seen:
                order.append(name)
                seen.add(name)
        for name in sheets.keys():
            if name not in seen:
                order.append(name)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for name in order:
            df = sheets[name]
            sheet = name[:31]

            df.to_excel(writer, sheet_name=sheet, index=False)

            ws = writer.book[sheet]
            _autosize(ws, df)
            _format_numbers(ws, df)

            # TYLKO ten arkusz: zawsze przecinki w średnich wymiarach
            if name == "package_type_share":
                _force_comma_text_for_columns(
                    ws,
                    ["avg_length", "avg_width", "avg_height"],
                    decimals=2,
                )

                # wpisz podsumowanie 
                if package_type_share_summary is not None:
                    wavg_len, pred_eff = package_type_share_summary
                    _write_package_type_share_summary(ws, wavg_len, pred_eff)

            if descriptions and name in descriptions:
                text, start_cell, end_cell = descriptions[name]
                _add_description_block(ws, text=text, start_cell=start_cell, end_cell=end_cell)
